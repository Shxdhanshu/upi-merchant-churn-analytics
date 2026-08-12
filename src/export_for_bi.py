"""Export warehouse tables to CSV and a formatted Excel workbook for the BI layer.

Power BI and Tableau can both read DuckDB over ODBC, but the driver setup is
fiddly on Windows and adds nothing analytically. The aggregate tables are small
-- hundreds of rows, not millions -- so exporting them is the pragmatic path:
the heavy lifting stays in SQL where it belongs, and the BI tool consumes
finished marts.

That split is deliberate and worth stating in an interview: dashboards should
consume modelled tables, not raw transactions. A BI tool doing its own joins and
aggregations over 7.4M rows is how you end up with four dashboards that each
report a different number for the same metric.

Outputs
-------
reports/bi_exports/*.csv          one per mart, for Power BI / Tableau
reports/bi_exports/network_analytics.xlsx  multi-sheet workbook with live formulas
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from warehouse import connect

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "reports" / "bi_exports"

# Each mart, with the question it answers. The comment becomes the sheet legend.
MARTS = {
    "network_health_monthly": "Monthly active merchants, TPV, success rate, merchant flows",
    "cohort_retention": "Cohort x month-index retention triangle (logo and TPV)",
    "bank_month_sr": "Issuer x month success rate and volume share",
    "sr_contribution": "Rate effect vs mix effect decomposition by issuer",
    "revenue_at_risk": "Run-rate revenue by merchant status",
    "realised_failure_loss": "Revenue lost to failed transactions, by month",
}
VIEWS = {
    "incident_culprit_ranking": "Issuers ranked by worst rate vs by network impact",
    "exposure_vs_churn": "Churn rate by issuer-exposure decile",
    "cohort_m3_retention": "Month-3 retention by cohort",
}

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=12)


def autosize(ws, df: pd.DataFrame, start_row: int) -> None:
    for i, col in enumerate(df.columns, start=1):
        width = max(len(str(col)), *(len(str(v)) for v in df[col].head(200))) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(max(width, 10), 30)


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    con = connect(read_only=True)

    frames: dict[str, pd.DataFrame] = {}
    for name in list(MARTS) + list(VIEWS):
        df = con.execute(f"SELECT * FROM {name}").df()
        for c in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[c]):
                df[c] = df[c].dt.strftime("%Y-%m-%d")
        df.to_csv(OUT / f"{name}.csv", index=False)
        frames[name] = df
        print(f"  {name:<32} {len(df):>6,} rows -> {name}.csv")
    con.close()

    # ---- Excel workbook ---------------------------------------------------
    xlsx = OUT / "network_analytics.xlsx"
    descriptions = {**MARTS, **VIEWS}
    with pd.ExcelWriter(xlsx, engine="openpyxl") as xw:
        pd.DataFrame({"Sheet": [], "Contents": []}).to_excel(xw, sheet_name="README", index=False)
        for name, df in frames.items():
            df.to_excel(xw, sheet_name=name[:31], index=False, startrow=2)

    wb = load_workbook(xlsx)

    # README sheet: index plus a couple of live cross-sheet formulas so the
    # workbook recalculates rather than shipping frozen numbers.
    ws = wb["README"]
    for row in ws["A1:B2"]:
        for cell in row:
            cell.value = None
    ws["A1"] = "UPI Merchant Retention -- analytical marts"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Generated from data/warehouse.duckdb by src/export_for_bi.py. Do not edit sheet data by hand."
    ws["A2"].font = Font(name="Arial", italic=True, size=9)
    ws["A4"], ws["B4"], ws["C4"] = "Sheet", "Contents", "Rows"
    for c in ("A4", "B4", "C4"):
        ws[c].fill, ws[c].font = HEADER_FILL, HEADER_FONT
    for i, (name, desc) in enumerate(descriptions.items(), start=5):
        ws[f"A{i}"], ws[f"B{i}"] = name, desc
        # Sheets carry a title row, a description row and a header row above
        # the data, so the offset is 3. Getting this wrong still recalculates
        # cleanly -- it just reports the wrong count, which is worse.
        ws[f"C{i}"] = f"=COUNTA('{name[:31]}'!A:A)-3"
        for c in ("A", "B", "C"):
            ws[f"{c}{i}"].font = BODY_FONT

    r = 5 + len(descriptions) + 1
    ws[f"A{r}"] = "Headline metrics (live formulas)"
    ws[f"A{r}"].font = TITLE_FONT
    nh = "network_health_monthly"
    hdr = list(frames[nh].columns)
    col_sr = get_column_letter(hdr.index("success_rate") + 1)
    col_tpv = get_column_letter(hdr.index("tpv_inr") + 1)
    col_act = get_column_letter(hdr.index("active_merchants") + 1)
    n = len(frames[nh]) + 3
    checks = [
        ("Total TPV (INR)", f"=SUM('{nh}'!{col_tpv}4:{col_tpv}{n})", "#,##0"),
        ("Average success rate", f"=AVERAGE('{nh}'!{col_sr}4:{col_sr}{n})", "0.00%"),
        ("Lowest monthly success rate", f"=MIN('{nh}'!{col_sr}4:{col_sr}{n})", "0.00%"),
        ("Peak active merchants", f"=MAX('{nh}'!{col_act}4:{col_act}{n})", "#,##0"),
        ("Months below 92% success rate",
         f"=COUNTIF('{nh}'!{col_sr}4:{col_sr}{n},\"<0.92\")", "#,##0"),
    ]
    for j, (label, formula, fmt) in enumerate(checks, start=r + 1):
        ws[f"A{j}"], ws[f"B{j}"] = label, formula
        ws[f"A{j}"].font = BODY_FONT
        ws[f"B{j}"].font = BODY_FONT
        ws[f"B{j}"].number_format = fmt
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 10

    # Format every data sheet.
    for name, df in frames.items():
        s = wb[name[:31]]
        s["A1"] = name
        s["A1"].font = TITLE_FONT
        s["A2"] = descriptions[name]
        s["A2"].font = Font(name="Arial", italic=True, size=9)
        for cell in s[3]:
            if cell.value is not None:
                cell.fill, cell.font = HEADER_FILL, HEADER_FONT
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
        for row in s.iter_rows(min_row=4):
            for cell in row:
                cell.font = BODY_FONT
        s.freeze_panes = "A4"
        autosize(s, df, 3)

    wb.save(xlsx)
    print(f"\n  Excel workbook -> {xlsx.relative_to(ROOT)}")
    print(f"  {len(frames)} sheets, formulas on README")


if __name__ == "__main__":
    main()
